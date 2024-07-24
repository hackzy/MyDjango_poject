from openpyxl import Workbook
from django.db.models import Q
#=======================================================
from zcglxt.models import Data_All,Departments,Edit_Log
from zcglxt.general_def import get_edit_data
class ExportExcel:
    
    def __init__(self,workbook:Workbook) -> None:
        self.wb = workbook
        
    def set_worksheet(self,start_date,end_date,worksheet_name,depart):
        end_date += " 23:59:59"
        ws = self.wb.worksheets[worksheet_name]
        log_data = Edit_Log.objects.filter(edit_date__gte=start_date,edit_date__lte=end_date)
        old_list = get_edit_data(log_data,depart)
        version = Edit_Log.objects.filter(Q(old_depart = depart) | Q(new_depart = depart)).latest('edit_date').edit_date.strftime('%Y%m%d')
        depart_cell = ws['A2']
        version_cell = ws['F2']
        depart_cell.value += depart
        version_cell.value += version
        data = Data_All.objects.filter(depart_name = Departments.objects.get(name = depart))
        index = '=row()-3'
        for value in data:
            ws.append([index,
                    value.number,
                    value.type_name.name,
                    value.model,
                    value.pos,
                    value.ip,
                    value.descr])
        if old_list != []:
            ws.append([""]) #插入2空行
            ws.append([""])
            last_row = ws.max_row
            for export in old_list:
                    ws.append([
                        "=row()-%d"%(last_row),
                        export.number,
                        export.type_name.name,
                        export.model,
                        export.pos,
                        export.ip,
                        export.descr,
                    ])
        last_row = ws.max_row
        foot_row = str(last_row + 1)
        ws['B' + foot_row].value = "机构复核："
        ws['D' + foot_row].value = "信科复核："
        ws['F' + foot_row].value = "日期："
